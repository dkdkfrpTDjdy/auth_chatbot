# -*- coding: utf-8 -*-
"""
IAM 최신 요청 파일 전처리 스크립트 (v5, fresh-only)

주요 변경점
1) 기존 index / by_team 산출물은 참고하지 않고, 최신 IAM 요청 파일만 기준으로 전처리
2) auth_name == 'PRS' 인 데이터는 전 구간에서 제거
3) 결과 JSON 산출물(index / by_team)은 기존 파일을 정리한 뒤 전부 다시 생성
   - source 에 없는 팀이 남지 않도록 role_bundle_team_*.jsonl 도 정리
4) 기존 매핑 워크북은 구조/레벨 lookup 용도로만 사용
   - 과거 팀/권한 데이터를 보존하거나 merge 하지 않음
"""

from __future__ import annotations

import json
import re
import shutil
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


EXCEL_FILE_CACHE: Dict[str, pd.ExcelFile] = {}
SHEET_DF_CACHE: Dict[Tuple[str, str], pd.DataFrame] = {}
WB_CACHE: Dict[str, object] = {}
MAPPING_DETAIL_CACHE: Dict[Tuple[str, str], pd.DataFrame] = {}


CONFIG: Dict = {
    "paths": {
        # 최신 IAM 요청 원본
        "source_xlsx": r"D:\works\GEN_AI\auth_chat\auth_chat_2_restore\(SR90000020809)IAM 내 권한 최신 데이터 요청.xlsx",
        # 기존 매핑 워크북(lookup + 템플릿)
        "mapping_workbook": r"D:\works\GEN_AI\auth_chat\auth_chat_2_restore\사용자_조직_권한_메뉴 매핑_20251218_v1.2.xlsx",

        # 결과물
        "output_workbook":r"D:\works\GEN_AI\auth_chat\auth_chat_2_restore\사용자_조직_권한_메뉴 매핑_20251218_v1.2_updated_v5.xlsx",
        "output_debug_xlsx": r"D:\works\GEN_AI\auth_chat\auth_chat_2_restore\OUTPUT_팀별권한_통합_결과_v5.xlsx",
        "out_base": r"D:\works\GEN_AI\auth_chat\auth_chat_2_restore\public_data_v5",
    },
    "source_sheets": {
        "sap_users": ["SAP 권한별 임직원"],
        "ias_users": ["IAS 권한별 임직원"],
        "mro_users": ["MRO 권한별 임직원"],
        "srm_users": ["SRM 권한별 임직원"],
        "eaccount_users": ["eAccount 권한별 임직원"],
        "team_target": ["팀별 권한"],
        "sap_role_tcode": ["SAP 역할별 TCODE"],
        "role_menu": ["역할별 메뉴"],
    },
    "mapping_sheets": {
        "ias_user_detail": ["IAS_Sales"],
        "ias_team_detail": ["IAS_sales_조직"],
        "sap_user_detail": ["SAP"],
    },
    "user_cols": {
        "name": ["이름"],
        "empno": ["사번"],
        "sys_name": ["시스템명"],
        "sys_code": ["시스템코드"],
        "role_name": ["역할명"],
        "role_code": ["역할코드"],
        "desc": ["설명"],
        "start_date": ["시작일자"],
        "end_date": ["종료일자"],
        "dept_name": ["부서명"],
        "dept_code": ["부서코드"],
    },
    "team_target_cols": {
        "team_name": ["팀명"],
        "team_code": ["팀코드"],
        "role_name": ["역할명"],
        "sys_name": ["리소스"],
        "sys_code": ["리소스코드"],
        "start_date": ["시작일자"],
        "end_date": ["종료일자"],
    },
    "role_menu_cols": {
        "role_id": ["역할ID"],
        "role_name": ["역할명"],
        "menu_id": ["메뉴ID"],
        "menu_name": ["메뉴명"],
        "url": ["URL"],
    },
    "sap_tcode_cols": {
        "role_name": ["역할명"],
        "role_code": ["역할코드"],
        "menu_name": ["메뉴명"],
        "menu_code": ["메뉴코드"],
    },
    "constants": {
        "sap_desc_topn": 3,
        "log_sheet": "전처리_LOG",
        "debug_sheet1": "팀역할_통합",
        "debug_sheet2": "팀역할_통합_메뉴",
        "debug_sheet3": "로그",
        "compare_report_file": "compare_report.txt",
    },
}


# =========================
# 기본 유틸
# =========================
def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [str(c).strip() for c in out.columns]
    return out


def norm_text(x) -> str:
    if x is None or pd.isna(x):
        return ""
    s = str(x)
    if "\r" in s:
        s = s.replace("\r\n", "\n").replace("\r", "\n")
    if "\n" in s:
        s = "\n".join(part.rstrip(" \t") for part in s.split("\n"))
    return s.strip()


def norm_code(x) -> str:
    if x is None or pd.isna(x):
        return ""
    if isinstance(x, int):
        return str(x)
    if isinstance(x, float):
        if float(x).is_integer():
            return str(int(x))
        return str(x).rstrip("0").rstrip(".")
    s = str(x).strip()
    if re.fullmatch(r"-?\d+\.0", s):
        s = s[:-2]
    return s


def canon_team_code(x) -> str:
    s = norm_code(x)
    if not s:
        return ""
    if s.startswith("0RULE_"):
        return s
    if re.fullmatch(r"\d+", s):
        s2 = s.lstrip("0")
        return s2 if s2 else "0"
    return s


def pick_first_existing_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    for c in candidates:
        if c in df.columns:
            return c
    return None


def ensure_any_col(df: pd.DataFrame, candidates: List[str], label: str, sheet_name: str) -> str:
    c = pick_first_existing_col(df, candidates)
    if c is None:
        raise ValueError(f"[{sheet_name}] '{label}' 컬럼 후보가 없습니다: {candidates}\n현재 컬럼: {list(df.columns)}")
    return c


def get_excel_file(path: Path) -> pd.ExcelFile:
    key = str(path)
    if key not in EXCEL_FILE_CACHE:
        EXCEL_FILE_CACHE[key] = pd.ExcelFile(path)
    return EXCEL_FILE_CACHE[key]


def resolve_sheet_name(path: Path, candidates: List[str]) -> str:
    xls = get_excel_file(path)
    for c in candidates:
        if c in xls.sheet_names:
            return c
    raise ValueError(f"시트를 찾지 못했습니다. 후보={candidates}, 실제={xls.sheet_names}")


def read_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    key = (str(path), sheet_name)
    if key not in SHEET_DF_CACHE:
        xls = get_excel_file(path)
        SHEET_DF_CACHE[key] = clean_columns(xls.parse(sheet_name=sheet_name))
    return SHEET_DF_CACHE[key].copy()


def get_openpyxl_workbook(path: Path):
    key = str(path)
    if key not in WB_CACHE:
        WB_CACHE[key] = load_workbook(filename=path, read_only=True, data_only=True)
    return WB_CACHE[key]


def read_sheet_selected(path: Path, sheet_name: str, selected_columns: List[str]) -> pd.DataFrame:
    """
    대형 시트에서 필요한 컬럼만 스트리밍으로 읽는다.
    pandas.read_excel/openpyxl 전체 파싱보다 훨씬 가볍게 동작하도록 사용.
    """
    wb = get_openpyxl_workbook(path)
    ws = wb[sheet_name]

    rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    header_row = next(rows, None)
    if header_row is None:
        return pd.DataFrame(columns=selected_columns)

    header = [norm_text(x) for x in header_row]
    idx_map = {name: idx for idx, name in enumerate(header) if name}

    use_cols = []
    for c in selected_columns:
        if c in idx_map:
            use_cols.append((c, idx_map[c]))

    if not use_cols:
        return pd.DataFrame(columns=selected_columns)

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {}
        has_value = False
        for col_name, idx in use_cols:
            val = row[idx] if idx < len(row) else None
            rec[col_name] = val
            if val is not None and not (isinstance(val, str) and val.strip() == ''):
                has_value = True
        if has_value:
            records.append(rec)

    df = pd.DataFrame(records)
    for c in selected_columns:
        if c not in df.columns:
            df[c] = pd.NA
    return clean_columns(df[selected_columns])


def load_mapping_detail_sheet(path_mapping: Path, sheet_name: str) -> pd.DataFrame:
    key = (str(path_mapping), sheet_name)
    if key not in MAPPING_DETAIL_CACHE:
        selected = [
            'sys_code', 'sys_name', 'auth_name', 'auth_code', 'auth_desc',
            'menu_id', '1level', '2level', '3level'
        ]
        MAPPING_DETAIL_CACHE[key] = read_sheet_selected(path_mapping, sheet_name, selected)
    return MAPPING_DETAIL_CACHE[key].copy()


def pick_mode_non_empty(series: pd.Series) -> str:
    values = [norm_text(v) for v in series.tolist() if norm_text(v)]
    if not values:
        return ""
    return Counter(values).most_common(1)[0][0]


def build_desc_from_names(names: List[str], topn: int = 3) -> str:
    uniq: List[str] = []
    seen = set()
    for x in names:
        x = norm_text(x)
        if not x or x in seen:
            continue
        seen.add(x)
        uniq.append(x)
    if not uniq:
        return ""
    return f"{', '.join(uniq[:topn])} 등이 있습니다."


def first_non_empty(series: pd.Series, is_code: bool = False) -> str:
    for v in series.tolist():
        vv = norm_code(v) if is_code else norm_text(v)
        if vv:
            return vv
    return ""


def preferred_team_name(names: Iterable[str]) -> str:
    values = [norm_text(x) for x in names if norm_text(x)]
    if not values:
        return ""
    # '_' 포함 우선, 그 다음 길이가 긴 값 우선
    values = sorted(set(values), key=lambda x: ("_" not in x, -len(x), x))
    return values[0]


def is_prs_name(x) -> bool:
    return norm_text(x).upper() == "PRS"


def remove_prs_rows(df: pd.DataFrame, auth_col: str = "auth_name") -> Tuple[pd.DataFrame, int]:
    if auth_col not in df.columns:
        return df.copy(), 0
    mask = df[auth_col].map(is_prs_name)
    removed = int(mask.sum())
    return df.loc[~mask].copy(), removed


# =========================
# source canonicalize
# =========================
def load_user_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    df = read_sheet(path, sheet_name)
    c = CONFIG["user_cols"]

    c_name = ensure_any_col(df, c["name"], "이름", sheet_name)
    c_emp = ensure_any_col(df, c["empno"], "사번", sheet_name)
    c_sys_name = ensure_any_col(df, c["sys_name"], "시스템명", sheet_name)
    c_sys_code = pick_first_existing_col(df, c["sys_code"])
    c_role_name = ensure_any_col(df, c["role_name"], "역할명", sheet_name)
    c_role_code = ensure_any_col(df, c["role_code"], "역할코드", sheet_name)
    c_desc = pick_first_existing_col(df, c["desc"])
    c_start = pick_first_existing_col(df, c["start_date"])
    c_end = pick_first_existing_col(df, c["end_date"])
    c_dept_name = ensure_any_col(df, c["dept_name"], "부서명", sheet_name)
    c_dept_code = ensure_any_col(df, c["dept_code"], "부서코드", sheet_name)

    out = pd.DataFrame({
        "user_name": df[c_name].map(norm_text),
        "user_id": df[c_emp].map(norm_code),
        "sys_name": df[c_sys_name].map(norm_text),
        "sys_code": df[c_sys_code].map(norm_text) if c_sys_code else df[c_sys_name].map(norm_text),
        "auth_name": df[c_role_name].map(norm_text),
        "auth_code": df[c_role_code].map(norm_code),
        "auth_desc": df[c_desc].map(norm_text) if c_desc else "",
        "start_date": df[c_start] if c_start else pd.NaT,
        "end_date": df[c_end] if c_end else pd.NaT,
        "team_name": df[c_dept_name].map(norm_text),
        "team_code": df[c_dept_code].map(canon_team_code),
    })
    return out


def load_team_target_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    df = read_sheet(path, sheet_name)
    c = CONFIG["team_target_cols"]

    c_team_name = ensure_any_col(df, c["team_name"], "팀명", sheet_name)
    c_team_code = ensure_any_col(df, c["team_code"], "팀코드", sheet_name)
    c_role_name = ensure_any_col(df, c["role_name"], "역할명", sheet_name)
    c_sys_name = ensure_any_col(df, c["sys_name"], "리소스", sheet_name)
    c_sys_code = ensure_any_col(df, c["sys_code"], "리소스코드", sheet_name)
    c_start = pick_first_existing_col(df, c["start_date"])
    c_end = pick_first_existing_col(df, c["end_date"])

    out = pd.DataFrame({
        "team_name": df[c_team_name].map(norm_text),
        "team_code": df[c_team_code].map(canon_team_code),
        "auth_name": df[c_role_name].map(norm_text),
        "sys_name": df[c_sys_name].map(norm_text),
        "sys_code": df[c_sys_code].map(norm_text),
        "start_date": df[c_start] if c_start else pd.NaT,
        "end_date": df[c_end] if c_end else pd.NaT,
    }).drop_duplicates()
    return out


def load_role_menu(path: Path, sheet_name: str) -> pd.DataFrame:
    df = read_sheet(path, sheet_name)
    c = CONFIG["role_menu_cols"]
    return pd.DataFrame({
        "role_id": df[ensure_any_col(df, c["role_id"], "역할ID", sheet_name)].map(norm_code),
        "auth_name": df[ensure_any_col(df, c["role_name"], "역할명", sheet_name)].map(norm_text),
        "menu_id": df[ensure_any_col(df, c["menu_id"], "메뉴ID", sheet_name)].map(norm_code),
        "menu_name": df[ensure_any_col(df, c["menu_name"], "메뉴명", sheet_name)].map(norm_text),
        "url": df[ensure_any_col(df, c["url"], "URL", sheet_name)].map(norm_text),
    }).drop_duplicates()


def load_sap_tcode(path: Path, sheet_name: str) -> pd.DataFrame:
    df = read_sheet(path, sheet_name)
    c = CONFIG["sap_tcode_cols"]
    return pd.DataFrame({
        "auth_name": df[ensure_any_col(df, c["role_name"], "역할명", sheet_name)].map(norm_text),
        "auth_code": df[ensure_any_col(df, c["role_code"], "역할코드", sheet_name)].map(norm_code),
        "menu_name": df[ensure_any_col(df, c["menu_name"], "메뉴명", sheet_name)].map(norm_text),
        "menu_id": df[ensure_any_col(df, c["menu_code"], "메뉴코드", sheet_name)].map(norm_code),
    }).drop_duplicates()


# =========================
# legacy compare / team alias
# =========================
def find_first_existing_file(base: Path, candidates: List[str]) -> Optional[Path]:
    for name in candidates:
        p = base / name
        if p.exists():
            return p
        if p.suffix == "":
            p_json = base / f"{name}.json"
            if p_json.exists():
                return p_json
    return None


def load_json_flexible(path: Path):
    text = path.read_text(encoding="utf-8")
    return json.loads(text)


def load_legacy_compare_data(base: Path) -> Dict:
    result = {
        "teams_records": [],
        "systems_by_team": {},
        "roles_by_team_sys": {},
        "team_name_map": {},
        "found": False,
        "files": {},
    }
    if not base.exists():
        return result

    p_teams = find_first_existing_file(base, ["index_teams_V1", "index_teams"])
    p_systems = find_first_existing_file(base, ["index_systems_by_team"])
    p_roles = find_first_existing_file(base, ["index_roles_by_team_sys"])

    if p_teams:
        raw = load_json_flexible(p_teams)
        teams = raw.get("teams", raw) if isinstance(raw, dict) else raw
        if isinstance(teams, list):
            result["teams_records"] = teams
            team_map = defaultdict(list)
            for row in teams:
                tc = canon_team_code(row.get("team_code", ""))
                tn = norm_text(row.get("team_name", ""))
                if tc and tn:
                    team_map[tc].append(tn)
            result["team_name_map"] = {tc: preferred_team_name(names) for tc, names in team_map.items()}
        result["files"]["teams"] = str(p_teams)

    if p_systems:
        raw = load_json_flexible(p_systems)
        if isinstance(raw, dict):
            result["systems_by_team"] = raw
        result["files"]["systems"] = str(p_systems)

    if p_roles:
        raw = load_json_flexible(p_roles)
        if isinstance(raw, dict):
            result["roles_by_team_sys"] = raw
        result["files"]["roles"] = str(p_roles)

    result["found"] = bool(p_teams or p_systems or p_roles)
    return result


def enrich_team_name_with_legacy(df: pd.DataFrame, legacy_team_name_map: Dict[str, str], team_code_col: str = "team_code", team_name_col: str = "team_name") -> pd.DataFrame:
    out = df.copy()
    if team_code_col not in out.columns or team_name_col not in out.columns:
        return out

    def _pick(row) -> str:
        tc = canon_team_code(row.get(team_code_col, ""))
        old_name = legacy_team_name_map.get(tc, "")
        new_name = norm_text(row.get(team_name_col, ""))
        return old_name or new_name

    out[team_name_col] = out.apply(_pick, axis=1)
    return out


# =========================
# lookup (기존 매핑 워크북 활용)
# =========================
def build_ias_level_lookup(path_mapping: Path) -> pd.DataFrame:
    frames = []
    for key in ["ias_user_detail", "ias_team_detail"]:
        sh = resolve_sheet_name(path_mapping, CONFIG["mapping_sheets"][key])
        df = load_mapping_detail_sheet(path_mapping, sh)
        cols = {x: pick_first_existing_col(df, [x]) for x in ["menu_id", "1level", "2level", "3level"]}
        if not all(cols.values()):
            continue
        frames.append(pd.DataFrame({
            "menu_id": df[cols["menu_id"]].map(norm_code),
            "menu_name": df[cols["3level"]].map(norm_text),
            "1level": df[cols["1level"]].map(norm_text),
            "2level": df[cols["2level"]].map(norm_text),
            "3level": df[cols["3level"]].map(norm_text),
        }))
    all_df = pd.concat(frames, ignore_index=True).drop_duplicates()
    return (
        all_df.groupby(["menu_id", "menu_name"], dropna=False)
        .agg({"1level": pick_mode_non_empty, "2level": pick_mode_non_empty, "3level": pick_mode_non_empty})
        .reset_index()
    )


def build_sap_level_lookup(path_mapping: Path) -> pd.DataFrame:
    sh = resolve_sheet_name(path_mapping, CONFIG["mapping_sheets"]["sap_user_detail"])
    df = load_mapping_detail_sheet(path_mapping, sh)
    all_df = pd.DataFrame({
        "menu_id": df[ensure_any_col(df, ["menu_id"], "menu_id", sh)].map(norm_code),
        "menu_name": df[ensure_any_col(df, ["3level"], "3level", sh)].map(norm_text),
        "1level": df[ensure_any_col(df, ["1level"], "1level", sh)].map(norm_text),
        "2level": df[ensure_any_col(df, ["2level"], "2level", sh)].map(norm_text),
        "3level": df[ensure_any_col(df, ["3level"], "3level", sh)].map(norm_text),
    }).drop_duplicates()

    return (
        all_df.groupby(["menu_id", "menu_name"], dropna=False)
        .agg({"1level": pick_mode_non_empty, "2level": pick_mode_non_empty, "3level": pick_mode_non_empty})
        .reset_index()
    )


def build_old_role_catalog_from_mapping(path_mapping: Path) -> pd.DataFrame:
    frames = []
    for key in ["ias_user_detail", "ias_team_detail", "sap_user_detail"]:
        sh = resolve_sheet_name(path_mapping, CONFIG["mapping_sheets"][key])
        df = load_mapping_detail_sheet(path_mapping, sh)
        need = {"sys_code", "sys_name", "auth_name", "auth_code", "auth_desc"}
        if not need.issubset(set(df.columns)):
            continue
        frames.append(df[list(need)].copy())
    if not frames:
        return pd.DataFrame(columns=["sys_code", "sys_name", "auth_name", "auth_code", "auth_desc"])
    out = pd.concat(frames, ignore_index=True).drop_duplicates()
    for c in ["sys_code", "sys_name", "auth_name", "auth_code", "auth_desc"]:
        if c == "auth_code":
            out[c] = out[c].map(norm_code)
        else:
            out[c] = out[c].map(norm_text)
    out, _ = remove_prs_rows(out, "auth_name")
    return out.drop_duplicates()


# =========================
# catalog / mapping helpers
# =========================
def fill_sap_auth_desc(users_all: pd.DataFrame, sap_tcode: pd.DataFrame) -> pd.DataFrame:
    out = users_all.copy()
    sap_desc_map = (
        sap_tcode.groupby("auth_code")["menu_name"]
        .apply(lambda s: build_desc_from_names(s.tolist(), topn=int(CONFIG["constants"]["sap_desc_topn"])))
        .to_dict()
    )
    mask = (out["sys_code"] == "SAP") & out["auth_desc"].map(norm_text).eq("")
    out.loc[mask, "auth_desc"] = out.loc[mask, "auth_code"].map(lambda x: sap_desc_map.get(norm_code(x), ""))
    return out


def build_role_catalog(users_all: pd.DataFrame, old_catalog: pd.DataFrame) -> pd.DataFrame:
    now_cat = users_all[["sys_code", "sys_name", "auth_name", "auth_code", "auth_desc"]].copy().drop_duplicates()
    out = pd.concat([now_cat, old_catalog], ignore_index=True).drop_duplicates()
    out["sys_code"] = out["sys_code"].map(norm_text)
    out["sys_name"] = out["sys_name"].map(norm_text)
    out["auth_name"] = out["auth_name"].map(norm_text)
    out["auth_code"] = out["auth_code"].map(norm_code)
    out["auth_desc"] = out["auth_desc"].map(norm_text)
    out, _ = remove_prs_rows(out, "auth_name")
    return out.drop_duplicates()


def role_catalog_maps(role_catalog: pd.DataFrame) -> Tuple[Dict[Tuple[str, str], Dict[str, str]], Dict[Tuple[str, str], Dict[str, str]]]:
    by_sys_role: Dict[Tuple[str, str], Dict[str, str]] = {}
    by_sys_code: Dict[Tuple[str, str], Dict[str, str]] = {}

    for _, r in role_catalog.iterrows():
        key1 = (norm_text(r["sys_code"]), norm_text(r["auth_name"]))
        cur1 = by_sys_role.get(key1, {})
        by_sys_role[key1] = {
            "sys_name": cur1.get("sys_name") or norm_text(r["sys_name"]),
            "auth_code": cur1.get("auth_code") or norm_code(r["auth_code"]),
            "auth_desc": cur1.get("auth_desc") or norm_text(r["auth_desc"]),
            "auth_name": norm_text(r["auth_name"]),
        }

        key2 = (norm_text(r["sys_code"]), norm_code(r["auth_code"]))
        if norm_code(r["auth_code"]):
            cur2 = by_sys_code.get(key2, {})
            by_sys_code[key2] = {
                "sys_name": cur2.get("sys_name") or norm_text(r["sys_name"]),
                "auth_code": norm_code(r["auth_code"]),
                "auth_desc": cur2.get("auth_desc") or norm_text(r["auth_desc"]),
                "auth_name": cur2.get("auth_name") or norm_text(r["auth_name"]),
            }

    return by_sys_role, by_sys_code


def build_role_menu_maps(role_menu: pd.DataFrame) -> Tuple[Dict[str, List[Dict]], Dict[str, List[Dict]]]:
    by_code: Dict[str, List[Dict]] = defaultdict(list)
    by_name: Dict[str, List[Dict]] = defaultdict(list)
    for _, r in role_menu.iterrows():
        rec = {
            "role_id": norm_code(r["role_id"]),
            "auth_name": norm_text(r["auth_name"]),
            "menu_id": norm_code(r["menu_id"]),
            "menu_name": norm_text(r["menu_name"]),
            "url": norm_text(r["url"]),
        }
        by_code[rec["role_id"]].append(rec)
        by_name[rec["auth_name"]].append(rec)
    return by_code, by_name


def build_sap_tcode_map(sap_tcode: pd.DataFrame) -> Dict[str, List[Dict]]:
    out: Dict[str, List[Dict]] = defaultdict(list)
    for _, r in sap_tcode.iterrows():
        out[norm_code(r["auth_code"])].append({
            "menu_id": norm_code(r["menu_id"]),
            "menu_name": norm_text(r["menu_name"]),
        })
    return out


# =========================
# expand functions
# =========================
def enrich_team_target(team_target: pd.DataFrame, by_sys_role: Dict[Tuple[str, str], Dict[str, str]]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    rows = []
    logs = []

    for _, r in team_target.iterrows():
        sys_code = norm_text(r["sys_code"])
        auth_name = norm_text(r["auth_name"])
        found = by_sys_role.get((sys_code, auth_name), None)

        auth_code = found["auth_code"] if found else ""
        auth_desc = found["auth_desc"] if found else ""
        sys_name = found["sys_name"] if found and found.get("sys_name") else norm_text(r["sys_name"])

        row = r.to_dict()
        row["sys_name"] = sys_name
        row["auth_code"] = auth_code
        row["auth_desc"] = auth_desc
        rows.append(row)

        if not auth_code:
            logs.append({
                "issue": "team_target auth_code missing",
                "sys_code": sys_code,
                "team_code": canon_team_code(r["team_code"]),
                "team_name": norm_text(r["team_name"]),
                "auth_name": auth_name,
            })

    out = pd.DataFrame(rows)
    out, removed = remove_prs_rows(out, "auth_name")
    if removed > 0:
        logs.append({"issue": "PRS removed from team_target", "removed_rows": removed})
    return out, pd.DataFrame(logs)


def expand_ias_with_role_menu(base_df: pd.DataFrame, role_menu_by_code: Dict[str, List[Dict]], role_menu_by_name: Dict[str, List[Dict]]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    rows = []
    logs = []

    for _, r in base_df.iterrows():
        auth_code = norm_code(r.get("auth_code", ""))
        auth_name = norm_text(r.get("auth_name", ""))

        menus = role_menu_by_code.get(auth_code) or role_menu_by_name.get(auth_name) or [None]
        matched = menus != [None]

        if not matched:
            logs.append({
                "issue": "IAS role_menu missing",
                "sys_code": norm_text(r.get("sys_code", "")),
                "team_code": canon_team_code(r.get("team_code", "")),
                "team_name": norm_text(r.get("team_name", "")),
                "auth_code": auth_code,
                "auth_name": auth_name,
            })

        for m in menus:
            row = r.to_dict()
            if m is None:
                row["menu_id"] = ""
                row["menu_name"] = ""
                row["url"] = ""
            else:
                row["menu_id"] = m["menu_id"]
                row["menu_name"] = m["menu_name"]
                row["url"] = m["url"]
            rows.append(row)

    out = pd.DataFrame(rows)
    out, removed = remove_prs_rows(out, "auth_name")
    if removed > 0:
        logs.append({"issue": "PRS removed after IAS role menu expand", "removed_rows": removed})
    return out, pd.DataFrame(logs)


def expand_sap_with_tcode(base_df: pd.DataFrame, sap_tcode_map: Dict[str, List[Dict]]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    rows = []
    logs = []

    for _, r in base_df.iterrows():
        auth_code = norm_code(r.get("auth_code", ""))
        menus = sap_tcode_map.get(auth_code) or [None]

        if menus == [None]:
            logs.append({
                "issue": "SAP tcode missing",
                "sys_code": norm_text(r.get("sys_code", "")),
                "team_code": canon_team_code(r.get("team_code", "")),
                "team_name": norm_text(r.get("team_name", "")),
                "auth_code": auth_code,
                "auth_name": norm_text(r.get("auth_name", "")),
            })

        for m in menus:
            row = r.to_dict()
            if m is None:
                row["menu_id"] = ""
                row["menu_name"] = ""
                row["url"] = ""
            else:
                row["menu_id"] = m["menu_id"]
                row["menu_name"] = m["menu_name"]
                row["url"] = ""
            rows.append(row)

    out = pd.DataFrame(rows)
    out, removed = remove_prs_rows(out, "auth_name")
    if removed > 0:
        logs.append({"issue": "PRS removed after SAP expand", "removed_rows": removed})
    return out, pd.DataFrame(logs)


def apply_level_lookup(df: pd.DataFrame, lookup_df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    out = df.copy()

    if "menu_id" not in out.columns:
        out["menu_id"] = ""
    if "menu_name" not in out.columns:
        out["menu_name"] = ""

    out["menu_id"] = out["menu_id"].map(norm_code)
    out["menu_name"] = out["menu_name"].map(norm_text)

    lookup = lookup_df.copy()
    lookup["menu_id"] = lookup["menu_id"].map(norm_code)
    lookup["menu_name"] = lookup["menu_name"].map(norm_text)

    out = out.merge(lookup, how="left", on=["menu_id", "menu_name"])
    for c in ["1level", "2level", "3level"]:
        out[c] = out[c].fillna("").map(norm_text)

    fail = out.loc[
        out["menu_id"].ne("") & out["3level"].eq(""),
        ["sys_code", "team_code", "team_name", "auth_code", "auth_name", "menu_id", "menu_name"]
    ].copy()
    fail["issue"] = "level lookup missing"
    return out, fail


# =========================
# user / team detail generation
# =========================
def generate_ias_user_detail(
    ias_users: pd.DataFrame,
    role_menu_by_code: Dict[str, List[Dict]],
    role_menu_by_name: Dict[str, List[Dict]],
    ias_level_lookup: pd.DataFrame,
) -> Tuple[pd.DataFrame, List[pd.DataFrame]]:
    base = ias_users.drop_duplicates(subset=["user_id", "team_code", "sys_code", "auth_code", "auth_name"]).copy()
    expanded, log_role = expand_ias_with_role_menu(base, role_menu_by_code, role_menu_by_name)
    leveled, log_level = apply_level_lookup(expanded, ias_level_lookup)

    cols = ["user_name", "user_id", "sys_name", "sys_code", "auth_name", "auth_code", "auth_desc",
            "end_date", "team_name", "team_code", "menu_id", "1level", "2level", "3level", "url"]
    for c in cols:
        if c not in leveled.columns:
            leveled[c] = ""
    out = leveled[cols].copy().rename(columns={"url": "URL"})
    out, _ = remove_prs_rows(out, "auth_name")
    return out, [log_role, log_level]


def generate_sap_user_detail(
    sap_users: pd.DataFrame,
    sap_tcode_map: Dict[str, List[Dict]],
    sap_level_lookup: pd.DataFrame,
) -> Tuple[pd.DataFrame, List[pd.DataFrame]]:
    base = sap_users.drop_duplicates(subset=["user_id", "team_code", "sys_code", "auth_code", "auth_name"]).copy()
    expanded, log_tcode = expand_sap_with_tcode(base, sap_tcode_map)
    leveled, log_level = apply_level_lookup(expanded, sap_level_lookup)

    cols = ["user_name", "user_id", "team_name", "team_code", "sys_name", "sys_code",
            "auth_name", "auth_code", "1level", "2level", "3level", "menu_id"]
    for c in cols:
        if c not in leveled.columns:
            leveled[c] = ""
    out = leveled[cols].copy()
    out["역할명2"] = out["auth_name"]
    out["역할 코드"] = out["auth_code"]
    out = out[["user_name", "user_id", "team_name", "team_code", "sys_name", "sys_code",
               "auth_name", "auth_code", "역할명2", "역할 코드", "1level", "2level", "3level", "menu_id"]]
    out, _ = remove_prs_rows(out, "auth_name")
    return out, [log_tcode, log_level]


def generate_ias_team_detail(
    team_target_lego: pd.DataFrame,
    role_menu_by_code: Dict[str, List[Dict]],
    role_menu_by_name: Dict[str, List[Dict]],
    ias_level_lookup: pd.DataFrame,
) -> Tuple[pd.DataFrame, List[pd.DataFrame]]:
    expanded, log_role = expand_ias_with_role_menu(team_target_lego, role_menu_by_code, role_menu_by_name)
    leveled, log_level = apply_level_lookup(expanded, ias_level_lookup)

    cols = ["auth_code", "auth_name", "auth_desc", "1level", "2level", "3level",
            "menu_id", "url", "end_date", "team_name", "team_code", "sys_name", "sys_code"]
    for c in cols:
        if c not in leveled.columns:
            leveled[c] = ""
    out = leveled[cols].copy().rename(columns={"url": "URL"})
    out, _ = remove_prs_rows(out, "auth_name")
    return out, [log_role, log_level]


# =========================
# JSON 산출물용 팀권한 통합
# =========================
def build_team_roles_from_users(users_all: pd.DataFrame) -> pd.DataFrame:
    out = (
        users_all[["team_name", "team_code", "sys_name", "sys_code", "auth_name", "auth_code", "auth_desc", "start_date", "end_date"]]
        .drop_duplicates(subset=["team_code", "sys_code", "auth_code", "auth_name"])
        .copy()
    )
    out, _ = remove_prs_rows(out, "auth_name")
    return out


def build_team_roles_union(team_target_enriched: pd.DataFrame, users_team_roles: pd.DataFrame) -> pd.DataFrame:
    target = team_target_enriched.copy()
    target["auth_code"] = target.apply(
        lambda r: norm_code(r["auth_code"]) if norm_code(r["auth_code"]) else f"NAME::{norm_text(r['auth_name'])}",
        axis=1,
    )
    target["auth_desc"] = target["auth_desc"].map(norm_text)

    union_df = pd.concat([
        target[["team_name", "team_code", "sys_name", "sys_code", "auth_name", "auth_code", "auth_desc", "start_date", "end_date"]],
        users_team_roles[["team_name", "team_code", "sys_name", "sys_code", "auth_name", "auth_code", "auth_desc", "start_date", "end_date"]],
    ], ignore_index=True)

    union_df, _ = remove_prs_rows(union_df, "auth_name")

    return (
        union_df.groupby(["team_code", "sys_code", "auth_code", "auth_name"], dropna=False)
        .agg(
            team_name=("team_name", first_non_empty),
            sys_name=("sys_name", first_non_empty),
            auth_desc=("auth_desc", first_non_empty),
            start_date=("start_date", "min"),
            end_date=("end_date", "max"),
        )
        .reset_index()
    )


def expand_team_roles_for_json(
    team_roles: pd.DataFrame,
    role_menu_by_code: Dict[str, List[Dict]],
    role_menu_by_name: Dict[str, List[Dict]],
    sap_tcode_map: Dict[str, List[Dict]],
    ias_level_lookup: pd.DataFrame,
    sap_level_lookup: pd.DataFrame,
) -> Tuple[pd.DataFrame, List[pd.DataFrame]]:
    logs: List[pd.DataFrame] = []

    is_lego = team_roles["sys_code"].eq("LEGO")
    is_sap = team_roles["sys_code"].eq("SAP")

    df_lego = team_roles[is_lego].copy()
    df_sap = team_roles[is_sap].copy()
    df_other = team_roles[~(is_lego | is_sap)].copy()

    lego_expanded, lego_log_role = expand_ias_with_role_menu(df_lego, role_menu_by_code, role_menu_by_name)
    lego_leveled, lego_log_level = apply_level_lookup(lego_expanded, ias_level_lookup)
    logs.extend([lego_log_role, lego_log_level])

    sap_expanded, sap_log_tcode = expand_sap_with_tcode(df_sap, sap_tcode_map)
    sap_leveled, sap_log_level = apply_level_lookup(sap_expanded, sap_level_lookup)
    logs.extend([sap_log_tcode, sap_log_level])

    for c in ["menu_id", "menu_name", "url", "1level", "2level", "3level"]:
        if c not in df_other.columns:
            df_other[c] = ""
    out = pd.concat([lego_leveled, sap_leveled, df_other], ignore_index=True, sort=False)

    for c in ["menu_id", "menu_name", "url", "1level", "2level", "3level"]:
        if c not in out.columns:
            out[c] = ""

    out, _ = remove_prs_rows(out, "auth_name")
    return out, logs


def build_role_meta_map(df: pd.DataFrame) -> Dict[Tuple[str, str, str], Dict[str, str]]:
    meta: Dict[Tuple[str, str, str], Dict[str, str]] = {}
    for (team_code, sys_code, auth_code), g in df.groupby(["team_code", "sys_code", "auth_code"]):
        meta[(team_code, sys_code, auth_code)] = {
            "auth_name": first_non_empty(g["auth_name"]),
            "auth_desc": first_non_empty(g["auth_desc"]),
        }
    return meta


def to_outputs(df: pd.DataFrame) -> Dict:
    out = df.copy()
    out, _ = remove_prs_rows(out, "auth_name")
    need_cols = ["team_code", "team_name", "sys_code", "sys_name", "auth_code", "auth_name", "auth_desc", "menu_id", "1level", "2level", "3level"]
    for c in need_cols:
        if c not in out.columns:
            out[c] = ""

    role_meta = build_role_meta_map(out)
    out["menu_id"] = out["menu_id"].map(norm_code)
    out["path"] = (
        out["1level"].map(norm_text) + " > " +
        out["2level"].map(norm_text) + " > " +
        out["3level"].map(norm_text)
    )

    teams_records = (
        out[["team_code", "team_name"]]
        .drop_duplicates()
        .sort_values(["team_name", "team_code"])
        .to_dict(orient="records")
    )

    systems_by_team: Dict[str, List[Dict[str, str]]] = {}
    for team_code, g in out.groupby("team_code"):
        sys_items = (
            g[["sys_code", "sys_name"]]
            .drop_duplicates()
            .sort_values(["sys_name", "sys_code"])
            .to_dict(orient="records")
        )
        systems_by_team[canon_team_code(team_code)] = [
            {"sys_code": norm_text(x["sys_code"]), "sys_name": norm_text(x["sys_name"])}
            for x in sys_items
        ]

    roles_by_team_sys: Dict[str, List[Dict[str, str]]] = {}
    for (team_code, sys_code), g in out.groupby(["team_code", "sys_code"]):
        roles = []
        for auth_code in sorted(g["auth_code"].map(norm_code).unique()):
            meta = role_meta.get((team_code, sys_code, auth_code), {"auth_name": "", "auth_desc": ""})
            if is_prs_name(meta["auth_name"]):
                continue
            roles.append({
                "auth_code": auth_code,
                "auth_name": meta["auth_name"],
                "auth_desc": meta["auth_desc"],
            })
        roles_by_team_sys[f"{canon_team_code(team_code)}|{norm_text(sys_code)}"] = sorted(
            roles, key=lambda x: (x["auth_name"], x["auth_code"])
        )

    bundles_by_team: Dict[str, Dict[str, Dict]] = {}
    for (team_code, sys_code, auth_code), g in out.groupby(["team_code", "sys_code", "auth_code"]):
        tc = canon_team_code(team_code)
        sys_code = norm_text(sys_code)
        auth_code = norm_code(auth_code)
        key = f"{sys_code}|{auth_code}"

        team_name = first_non_empty(g["team_name"])
        sys_name = first_non_empty(g["sys_name"])
        meta = role_meta.get((team_code, sys_code, auth_code), {"auth_name": "", "auth_desc": ""})
        if is_prs_name(meta["auth_name"]):
            continue

        menus = (
            g[["menu_id", "path"]]
            .drop_duplicates()
            .sort_values(["path", "menu_id"])
            .to_dict(orient="records")
        )
        bundles_by_team.setdefault(tc, {})
        bundles_by_team[tc][key] = {
            "team_code": tc,
            "team_name": team_name,
            "sys_code": sys_code,
            "sys_name": sys_name,
            "auth_code": auth_code,
            "auth_name": meta["auth_name"],
            "auth_desc": meta["auth_desc"],
            "menus": [{"menu_id": norm_code(m["menu_id"]), "path": norm_text(m["path"])} for m in menus],
        }

    return {
        "teams_records": teams_records,
        "systems_by_team": systems_by_team,
        "roles_by_team_sys": roles_by_team_sys,
        "bundles_by_team": bundles_by_team,
    }


# =========================
# compare report
# =========================
def normalize_team_records(teams_records: List[Dict]) -> Dict[str, str]:
    out: Dict[str, List[str]] = defaultdict(list)
    for row in teams_records:
        tc = canon_team_code(row.get("team_code", ""))
        tn = norm_text(row.get("team_name", ""))
        if tc and tn:
            out[tc].append(tn)
    return {tc: preferred_team_name(names) for tc, names in out.items()}


def normalize_systems_by_team(systems_by_team: Dict) -> Dict[str, set]:
    out = {}
    for team_code, rows in (systems_by_team or {}).items():
        tc = canon_team_code(team_code)
        items = set()
        for row in rows or []:
            items.add((norm_text(row.get("sys_code", "")), norm_text(row.get("sys_name", ""))))
        out[tc] = items
    return out


def normalize_roles_by_team_sys(roles_by_team_sys: Dict) -> Dict[str, set]:
    out = {}
    for key, rows in (roles_by_team_sys or {}).items():
        if "|" not in str(key):
            continue
        tc, sc = str(key).split("|", 1)
        k2 = f"{canon_team_code(tc)}|{norm_text(sc)}"
        items = set()
        for row in rows or []:
            auth_name = norm_text(row.get("auth_name", ""))
            if is_prs_name(auth_name):
                continue
            items.add((norm_code(row.get("auth_code", "")), auth_name, norm_text(row.get("auth_desc", ""))))
        out[k2] = items
    return out


def display_team(team_code: str, new_team_map: Dict[str, str], old_team_map: Dict[str, str]) -> str:
    name = new_team_map.get(team_code) or old_team_map.get(team_code) or ""
    return f"{name} ({team_code})" if name else f"{team_code}"


def build_compare_report(old_data: Dict, new_outputs: Dict) -> str:
    old_team_map = normalize_team_records(old_data.get("teams_records", []))
    new_team_map = normalize_team_records(new_outputs.get("teams_records", []))

    old_teams = set(old_team_map.keys())
    new_teams = set(new_team_map.keys())
    added_teams = sorted(new_teams - old_teams)
    removed_teams = sorted(old_teams - new_teams)
    common_teams = sorted(new_teams & old_teams)

    renamed = []
    for tc in common_teams:
        if norm_text(old_team_map.get(tc, "")) != norm_text(new_team_map.get(tc, "")):
            renamed.append((tc, old_team_map.get(tc, ""), new_team_map.get(tc, "")))

    old_sys = normalize_systems_by_team(old_data.get("systems_by_team", {}))
    new_sys = normalize_systems_by_team(new_outputs.get("systems_by_team", {}))
    sys_added_lines = []
    sys_removed_lines = []
    for tc in sorted(set(old_sys) | set(new_sys)):
        add = sorted(new_sys.get(tc, set()) - old_sys.get(tc, set()))
        rem = sorted(old_sys.get(tc, set()) - new_sys.get(tc, set()))
        if add:
            sys_added_lines.append(f"- {display_team(tc, new_team_map, old_team_map)}: {', '.join([f'{a[0]}:{a[1]}' for a in add])}")
        if rem:
            sys_removed_lines.append(f"- {display_team(tc, new_team_map, old_team_map)}: {', '.join([f'{a[0]}:{a[1]}' for a in rem])}")

    old_roles = normalize_roles_by_team_sys(old_data.get("roles_by_team_sys", {}))
    new_roles = normalize_roles_by_team_sys(new_outputs.get("roles_by_team_sys", {}))
    role_added_lines = []
    role_removed_lines = []
    for key in sorted(set(old_roles) | set(new_roles)):
        tc = key.split("|", 1)[0]
        sc = key.split("|", 1)[1]
        add = sorted(new_roles.get(key, set()) - old_roles.get(key, set()))
        rem = sorted(old_roles.get(key, set()) - new_roles.get(key, set()))
        if add:
            vals = [f"{a[0]}:{a[1]}" for a in add]
            role_added_lines.append(f"- {display_team(tc, new_team_map, old_team_map)} | {sc}: {', '.join(vals)}")
        if rem:
            vals = [f"{a[0]}:{a[1]}" for a in rem]
            role_removed_lines.append(f"- {display_team(tc, new_team_map, old_team_map)} | {sc}: {', '.join(vals)}")

    lines: List[str] = []
    lines.append("=" * 80)
    lines.append("[COMPARE] 기존 index 대비 신규 생성 결과 차이")
    lines.append("=" * 80)
    lines.append(f"기존 팀 수: {len(old_teams)}")
    lines.append(f"신규 팀 수: {len(new_teams)}")
    lines.append(f"추가된 팀 수: {len(added_teams)}")
    lines.append(f"제외된 팀 수: {len(removed_teams)}")
    lines.append(f"동일 팀코드 내 팀명 변경 수: {len(renamed)}")
    lines.append("")

    lines.append("[추가된 팀]")
    if added_teams:
        for tc in added_teams:
            lines.append(f"- {display_team(tc, new_team_map, old_team_map)}")
    else:
        lines.append("- 없음")
    lines.append("")

    lines.append("[제외된 팀]")
    if removed_teams:
        for tc in removed_teams:
            lines.append(f"- {display_team(tc, new_team_map, old_team_map)}")
    else:
        lines.append("- 없음")
    lines.append("")

    lines.append("[동일 팀코드, 팀명 변경]")
    if renamed:
        for tc, old_name, new_name in renamed:
            lines.append(f"- {tc}: '{old_name}' -> '{new_name}'")
    else:
        lines.append("- 없음")
    lines.append("")

    lines.append(f"[팀별 시스템 추가] {len(sys_added_lines)}건")
    lines.extend(sys_added_lines or ["- 없음"])
    lines.append("")

    lines.append(f"[팀별 시스템 제거] {len(sys_removed_lines)}건")
    lines.extend(sys_removed_lines or ["- 없음"])
    lines.append("")

    lines.append(f"[팀-시스템별 권한 추가] {len(role_added_lines)}건")
    lines.extend(role_added_lines or ["- 없음"])
    lines.append("")

    lines.append(f"[팀-시스템별 권한 제거] {len(role_removed_lines)}건")
    lines.extend(role_removed_lines or ["- 없음"])
    lines.append("")

    return "\n".join(lines)


# =========================
# write
# =========================
def write_debug_excel(path_out: Path, team_roles: pd.DataFrame, team_roles_menu: pd.DataFrame, logs: pd.DataFrame) -> None:
    path_out.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path_out, engine="openpyxl") as w:
        team_roles.to_excel(w, sheet_name=CONFIG["constants"]["debug_sheet1"][:31], index=False)
        team_roles_menu.to_excel(w, sheet_name=CONFIG["constants"]["debug_sheet2"][:31], index=False)
        logs.to_excel(w, sheet_name=CONFIG["constants"]["debug_sheet3"][:31], index=False)


def write_updated_workbook(template_path: Path, out_path: Path, sheet_map: Dict[str, pd.DataFrame]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, out_path)

    with pd.ExcelWriter(out_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
        for sheet_name, df in sheet_map.items():
            df.to_excel(w, sheet_name=sheet_name[:31], index=False)


def clear_managed_outputs(out_base: Path) -> None:
    """
    기존 산출물은 전부 무시하고 fresh-only 로 다시 생성한다.
    - index json 파일 삭제
    - by_team/role_bundle_team_*.jsonl 삭제
    다른 보조 파일은 건드리지 않는다.
    """
    out_base.mkdir(parents=True, exist_ok=True)
    for name in [
        "index_teams.json",
        "index_teams_V1.json",
        "index_systems_by_team.json",
        "index_roles_by_team_sys.json",
    ]:
        p = out_base / name
        if p.exists():
            p.unlink()

    by_team = out_base / "by_team"
    by_team.mkdir(parents=True, exist_ok=True)
    for p in by_team.glob("role_bundle_team_*.jsonl"):
        p.unlink()


def write_json_outputs(out_base: Path, outputs: Dict) -> None:
    clear_managed_outputs(out_base)
    by_team = out_base / "by_team"

    teams_payload = json.dumps({"teams": outputs["teams_records"]}, ensure_ascii=False, indent=2)
    (out_base / "index_teams.json").write_text(teams_payload, encoding="utf-8")
    # 운영 호환용으로 같이 생성
    (out_base / "index_teams_V1.json").write_text(teams_payload, encoding="utf-8")

    (out_base / "index_systems_by_team.json").write_text(
        json.dumps(outputs["systems_by_team"], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (out_base / "index_roles_by_team_sys.json").write_text(
        json.dumps(outputs["roles_by_team_sys"], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    for team_code, bundle_map in outputs["bundles_by_team"].items():
        p = by_team / f"role_bundle_team_{team_code}.jsonl"
        rows = sorted(
            bundle_map.values(),
            key=lambda x: (
                norm_text(x.get("sys_name", "")),
                norm_text(x.get("auth_name", "")),
                norm_code(x.get("auth_code", "")),
            ),
        )
        with p.open("w", encoding="utf-8") as f:
            for row in rows:
                f.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_compare_report(out_base: Path, report: str) -> Path:
    out_base.mkdir(parents=True, exist_ok=True)
    p = out_base / CONFIG["constants"]["compare_report_file"]
    p.write_text(report, encoding="utf-8")
    return p


# =========================
# main
# =========================
def main() -> None:
    source_path = Path(CONFIG["paths"]["source_xlsx"])
    mapping_path = Path(CONFIG["paths"]["mapping_workbook"])
    out_workbook = Path(CONFIG["paths"]["output_workbook"])
    out_debug = Path(CONFIG["paths"]["output_debug_xlsx"])
    out_base = Path(CONFIG["paths"]["out_base"])

    sh_sap_users = resolve_sheet_name(source_path, CONFIG["source_sheets"]["sap_users"])
    sh_ias_users = resolve_sheet_name(source_path, CONFIG["source_sheets"]["ias_users"])
    sh_mro_users = resolve_sheet_name(source_path, CONFIG["source_sheets"]["mro_users"])
    sh_srm_users = resolve_sheet_name(source_path, CONFIG["source_sheets"]["srm_users"])
    sh_eac_users = resolve_sheet_name(source_path, CONFIG["source_sheets"]["eaccount_users"])
    sh_team_target = resolve_sheet_name(source_path, CONFIG["source_sheets"]["team_target"])
    sh_sap_tcode = resolve_sheet_name(source_path, CONFIG["source_sheets"]["sap_role_tcode"])
    sh_role_menu = resolve_sheet_name(source_path, CONFIG["source_sheets"]["role_menu"])

    sh_out_ias = resolve_sheet_name(mapping_path, CONFIG["mapping_sheets"]["ias_user_detail"])
    sh_out_ias_org = resolve_sheet_name(mapping_path, CONFIG["mapping_sheets"]["ias_team_detail"])
    sh_out_sap = resolve_sheet_name(mapping_path, CONFIG["mapping_sheets"]["sap_user_detail"])

    sap_users = load_user_sheet(source_path, sh_sap_users)
    ias_users = load_user_sheet(source_path, sh_ias_users)
    mro_users = load_user_sheet(source_path, sh_mro_users)
    srm_users = load_user_sheet(source_path, sh_srm_users)
    eac_users = load_user_sheet(source_path, sh_eac_users)
    team_target = load_team_target_sheet(source_path, sh_team_target)
    role_menu = load_role_menu(source_path, sh_role_menu)
    sap_tcode = load_sap_tcode(source_path, sh_sap_tcode)

    removed_prs_total = 0
    loaded_frames = []
    for df in [sap_users, ias_users, mro_users, srm_users, eac_users]:
        dfx, removed = remove_prs_rows(df, "auth_name")
        removed_prs_total += removed
        loaded_frames.append(dfx)
    sap_users, ias_users, mro_users, srm_users, eac_users = loaded_frames
    team_target, removed = remove_prs_rows(team_target, "auth_name")
    removed_prs_total += removed

    users_all = pd.concat([sap_users, ias_users, mro_users, srm_users, eac_users], ignore_index=True)
    users_all = fill_sap_auth_desc(users_all, sap_tcode)
    users_all, removed = remove_prs_rows(users_all, "auth_name")
    removed_prs_total += removed

    ias_level_lookup = build_ias_level_lookup(mapping_path)
    sap_level_lookup = build_sap_level_lookup(mapping_path)

    # fresh-only: 과거 팀/권한 결과는 사용하지 않음.
    # 다만 최신 사용자 원천(users_all)만으로 role catalog 를 구성해 팀별 권한을 보강한다.
    empty_old_role_catalog = pd.DataFrame(columns=["sys_code", "sys_name", "auth_name", "auth_code", "auth_desc"])
    role_catalog = build_role_catalog(users_all, empty_old_role_catalog)
    by_sys_role, _ = role_catalog_maps(role_catalog)

    role_menu_by_code, role_menu_by_name = build_role_menu_maps(role_menu)
    sap_tcode_map = build_sap_tcode_map(sap_tcode)

    team_target_enriched, log_team_target = enrich_team_target(team_target, by_sys_role)

    df_ias_user_detail, logs_ias_user = generate_ias_user_detail(
        ias_users=users_all[users_all["sys_code"] == "LEGO"].copy(),
        role_menu_by_code=role_menu_by_code,
        role_menu_by_name=role_menu_by_name,
        ias_level_lookup=ias_level_lookup,
    )
    df_sap_user_detail, logs_sap_user = generate_sap_user_detail(
        sap_users=users_all[users_all["sys_code"] == "SAP"].copy(),
        sap_tcode_map=sap_tcode_map,
        sap_level_lookup=sap_level_lookup,
    )
    df_ias_team_detail, logs_ias_team = generate_ias_team_detail(
        team_target_lego=team_target_enriched[team_target_enriched["sys_code"] == "LEGO"].copy(),
        role_menu_by_code=role_menu_by_code,
        role_menu_by_name=role_menu_by_name,
        ias_level_lookup=ias_level_lookup,
    )

    users_team_roles = build_team_roles_from_users(users_all)
    team_roles = build_team_roles_union(team_target_enriched, users_team_roles)
    team_roles_menu, logs_team_json = expand_team_roles_for_json(
        team_roles=team_roles,
        role_menu_by_code=role_menu_by_code,
        role_menu_by_name=role_menu_by_name,
        sap_tcode_map=sap_tcode_map,
        ias_level_lookup=ias_level_lookup,
        sap_level_lookup=sap_level_lookup,
    )

    outputs_new = to_outputs(team_roles_menu)

    log_frames = [log_team_target] + logs_ias_user + logs_sap_user + logs_ias_team + logs_team_json
    log_frames = [x for x in log_frames if isinstance(x, pd.DataFrame) and len(x) > 0]
    if log_frames:
        df_log = pd.concat(log_frames, ignore_index=True, sort=False)
    else:
        df_log = pd.DataFrame([{"issue": "no issues"}])

    df_log = pd.concat([
        df_log,
        pd.DataFrame([{
            "issue": "PRS filtered summary",
            "removed_rows": removed_prs_total,
        }])
    ], ignore_index=True, sort=False)

    write_updated_workbook(mapping_path, out_workbook, {
        sh_out_ias: df_ias_user_detail,
        sh_out_ias_org: df_ias_team_detail,
        sh_out_sap: df_sap_user_detail,
        CONFIG["constants"]["log_sheet"]: df_log,
    })

    write_debug_excel(out_debug, team_roles, team_roles_menu, df_log)
    write_json_outputs(out_base, outputs_new)

    print("✅ 완료 (fresh-only)")
    print(f"- Updated workbook: {out_workbook}")
    print(f"- Debug workbook:   {out_debug}")
    print(f"- JSON base:        {out_base}")
    print()
    print("[정책] 기존 index / by_team 산출물은 무시하고 최신 IAM 요청 파일 기준으로 전부 다시 생성했습니다.")
    print("[정책] source 에 없는 팀은 결과에 남지 않아야 합니다.")
    print()
    print(f"- IAS_Sales rows:   {len(df_ias_user_detail)}")
    print(f"- IAS_sales_조직 rows: {len(df_ias_team_detail)}")
    print(f"- SAP rows:         {len(df_sap_user_detail)}")
    print(f"- Team roles rows:  {len(team_roles)}")
    print(f"- Team menu rows:   {len(team_roles_menu)}")
    print(f"- Teams count:      {len(outputs_new['teams_records'])}")
    print(f"- Log rows:         {len(df_log)}")
    print(f"- PRS removed rows: {removed_prs_total}")

if __name__ == "__main__":
    main()
